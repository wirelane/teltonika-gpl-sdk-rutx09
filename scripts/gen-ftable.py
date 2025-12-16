#!/usr/bin/python

from os import path
from subprocess import Popen, PIPE
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side

list_features = set()
list_targets = list()
list_profiles = list()

class Target:
    def __init__(self, name, features):
        self.name = name
        self.features = set(features.split())

class Profile:
    def __init__(self, name, target, features):
        self.name = name
        self.target = target
        self.features = set(features.split())
        self.features.update(target.features)

def parse_target(f, line):
    name = line[8:-1]

    for l in f:
        if '@@' in l:
            break
        elif 'Target-Features:' in l:
            features = l[17:-1]

    for i in features.split():
        list_features.add(i)

    t = Target(name, features)
    list_targets.append(t)

    return t

def parse_profile(f, target):
    for l in f:
        if '@@' in l:
            break
        elif 'Target-Profile-Name:' in l:
            name = l[21:-1]
        elif 'Target-Profile-Features:' in l:
            features = l[25:-1]

    for i in features.split():
        list_features.add(i)

    p = Profile(name, target, features)
    list_profiles.append(p)

def read_targetinfo():
    ti = 'tmp/.targetinfo'

    if not path.exists(ti):
        print('Target info does not exist, running defconfig...')
        Popen('make defconfig', shell=True).wait()

    if not path.exists(ti):
        print('Unable to generate ftable due to missing targetinfo data!')
        exit(1)

    f = open(ti, 'r')

    for line in f:
        if 'Target:' in line:
            last_target = parse_target(f, line)
        elif 'Target-Profile:' in line:
            parse_profile(f, last_target)

    f.close()

def fill_ftable():
    wb = Workbook()

    fill_dev_ftable(wb)
    fill_opt_ftable(wb)

    name = 'rutos-ftable.xlsx'
    wb.save(name)

    print(f'Generated {name} file')

def fill_dev_ftable(wb):
    ws = wb.active
    ws.title = 'Device List'
    ws.freeze_panes = 'B2'

    # add feature list
    c = 2
    for i in list_features:
        cell = ws.cell(row=1, column=c)
        cell.value = i
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        c += 1

    # add device list and select supported features
    r = 2
    for i in list_profiles:
        cell = ws.cell(row=r, column=1)
        cell.value = i.name
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

        for row in ws.iter_rows(min_row=1, min_col=1):
            for cell in row:
                for f in i.features:
                    if cell.value != f:
                        continue

                    c2 = ws.cell(row=r, column=cell.column)
                    colFill = PatternFill(start_color='295238', end_color='295238', fill_type='solid')
                    c2.fill = colFill

        r += 1

    resize_columns(ws)

def fill_opt_ftable(wb):
    ws = wb.create_sheet('Feature-Kconfig')
    ws.freeze_panes = ws['A2']

    cell = ws.cell(row=1, column=1)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    cell.value = 'Feature Name'

    cell = ws.cell(row=1, column=2)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    cell.value = 'Kconfig Option'

    r = 2

    for i in list_features:
        opt = ''

        with Popen(['scripts/target-metadata.pl', 'show', i], stdout=PIPE) as f:
            opt = f.stdout.read().rstrip()

        if not opt:
            continue

        cell = ws.cell(row=r, column=1)
        cell.alignment = Alignment(horizontal='center')
        cell.value = i

        cell = ws.cell(row=r, column=2)
        cell.alignment = Alignment(horizontal='center')
        cell.value = opt
        r += 1

    resize_columns(ws)

def resize_columns(ws):
    for c in ws.columns:
        max_len = 0
        column = c[0].column_letter

        for cell in c:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            try:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
            except:
                pass

        ws.column_dimensions[column].width = (max_len + 2) * 1.25

read_targetinfo()
list_features = sorted(list_features)
list_profiles = sorted(list_profiles, key=lambda i: i.features, reverse=True)
fill_ftable()
